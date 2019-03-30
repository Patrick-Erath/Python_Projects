# distance.py : This project uses Google Maps API to find the distance between a starting address
# and multiple destinations. The project finds the distance to travel between two places every HOUR
# The project then writes the distance (in hours/minutes) to each destination for a specific time to an excel file
# This was a personal project to determine the best time to leave for a destination

import googlemaps, datetime, openpyxl, time
from openpyxl.styles import Font
from datetime import timedelta

gmaps = googlemaps.Client(key='YOUR_API_KEY_HERE')
startingAddress = 'Schweinfurt, Germany'
destinations = ['Amsterdam', 'Berlin', 'Brussels','Stuttgart','Sächsische Schweiz','Allgaü','Garmish-Partenkirchen']
dictionary = {}
firstTime = True
currentRow = [2]

def getDistance():
	print('Getting distances using GoogleMaps API...')
	for i in range(len(destinations)):
		try:
			now = datetime.datetime.now()
			directions_result = gmaps.directions("Schweinfurt, Germany",
			                                     destinations[i],
			                                     mode="driving",
			                                     departure_time=now)
			# Populate dictionary
			dictionary[destinations[i]]=directions_result[0]['legs'][0]['duration']['text']
		except: 
			print('Error')

def addExcel():
	# Output to excel file
	print('Writing distances to Excel sheet...')
	if firstTime == True:
		wb = openpyxl.Workbook()
		sheet = wb.active
		sheet.title = 'Destination Times'
		# Init names
		fontObj1 = Font(name='Times New Roman', bold=True)
		sheet.cell(row=1, column=1).value='Destination'
		sheet.cell(row=1, column=2).value='Driving Time'
		sheet.cell(row=1, column=2).value='Time Taken'
		sheet['A1'].font=fontObj1
		sheet['B1'].font=fontObj1
		sheet['C1'].font=fontObj1
	else:
		wb = openpyxl.load_workbook('Time.xlsx')
		sheet = wb.active

	now = datetime.datetime.now()
	now = now.replace(microsecond=0)
	for destination in dictionary:
		rowNum = currentRow[0]
		sheet.cell(row=rowNum, column=1).value=destination
		sheet.cell(row=rowNum, column=2).value=dictionary[destination]
		sheet.cell(row=rowNum, column=3).value=now
		currentRow[0]+=1
	wb.save('Time.xlsx')
	print('Done')

def main():
	while 1:
		getDistance()
		addExcel()
		dt = datetime.datetime.now() + timedelta(seconds=10) # hours=1
		while datetime.datetime.now() < dt:
			time.sleep(1)
		global firstTime
		firstTime = False


if __name__ == "__main__":
	main()






