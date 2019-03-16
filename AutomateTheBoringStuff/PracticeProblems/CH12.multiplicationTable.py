import sys
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
sheet = wb.active

N = int(sys.argv[1])
bold = Font(bold=True)

# Initiliaze row and columns
for i in range(1,N+1):
	done = False
	#Populate Row Numbers (In Bold)
	sheet['A'+str(i+1)] = i
	sheet['A'+str(i+1)].font = bold

	#Populate Column Numbers (In Bold)
	sheet[get_column_letter(i+1)+str(1)] = i
	sheet[get_column_letter(i+1)+str(1)].font = bold

#TODO: Fill in Values
for row in range(1,N+1):
	for col in range(1,N+1):
		sheet[get_column_letter(col+1)+str(row+1)] = sheet[get_column_letter(col+1)+str(1)].value * sheet['A'+str(row+1)].value

wb.save('multiplicationTable.xlsx')