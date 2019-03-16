import sys
import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook('Invert.xlsx')
sheet = wb.active
wb_new = openpyxl.Workbook()
sheet_new = wb_new.active


for col in range(1, sheet.max_column+1):
	for row in range(1, sheet.max_row+1):
		#print(row, col)
		#print(sheet.cell(row=row, column = col).value)

		sheet_new.cell(row=col, column=row).value = sheet.cell(row=row, column = col).value

		#print(sheet.cell(row=row, column = col).value)
		#print(col,row)

		#print(sheet.cell(row=row, column = col).value)

wb_new.save('invertedGraph.xlsx')
