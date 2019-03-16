# blackRowInserter.py : inserts blank rows from row N to row M to spreadsheet
import sys
import openpyxl

# ENTER IN TERMINAL : blackRowInserter.py StartingRowNumber NumberOfBlankRows yourFile.xlsx
N = int(sys.argv[1]) # Starting row
M = int(sys.argv[2]) # Number of blank rows
filename = sys.argv[3] # File name

print('Opening Workbook...')
wb = openpyxl.load_workbook(filename)
sheet = wb.get_sheet_by_name('Sheet')
new_wb = openpyxl.Workbook()
new_sheet = new_wb.active

print('Adding Blank rows...')
for row in range(1, sheet.max_row+1):
	for col in range(1, sheet.max_column+1):
		if row < N:
			new_sheet.cell(row=row, column=col).value = sheet.cell(row=row, column=col).value
		else:
			new_sheet.cell(row=row+M, column=col).value = sheet.cell(row=row, column=col).value



print('Done')
new_wb.save('yesblank.xlsx')

