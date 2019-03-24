import openpyxl, os, csv, re

os.chdir(r'./Here')
path = os.getcwd()
os.makedirs('csv_files', exist_ok=True)
new_path = path + '\\' + 'csv_files'

for excelFile in os.listdir('.'):
	# Skip non-xlsx files, load the workbook object
	if not excelFile.endswith('.xlsx'):
		continue
	wb = openpyxl.load_workbook(excelFile)
	for sheetName in wb.sheetnames:
		# Loop through every sheet in the workbook
		sheet = wb[sheetName]
		# Create the CSV filename from the Excel filename and sheet title
		wbName = re.sub('.xlsx', '', excelFile) # Removing extension
		csvName = wbName + '_' + sheetName + '.csv' # Newly created csv name
		# Create the csv.writer object for this CSV file
		outputFile = open(csvName, 'w', newline='')
		outputWriter = csv.writer(outputFile)

		# Loop through every row in the sheet
		for rowNum in range(1, sheet.max_row + 1):
			rowData = []
			# Loop through every cell in the row
			for colNum in range(1, sheet.max_column + 1):
				# Append each cell's data to rowData
				cellData = sheet.cell(row = rowNum, column = colNum).value
				rowData.append(cellData)
			# Write the rowData list to the CSV file
			outputWriter.writerow(rowData)
		outputFile.close()
		shutil.move(os.path.join(path, csv_name), os.path.join(new_path, csv_name))